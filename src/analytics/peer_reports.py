import sqlite3
import pandas as pd
import os

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
DATABASE = "nifty100.db"

GREEN_FILL = PatternFill(
    fill_type="solid",
    start_color="90EE90",
    end_color="90EE90"
)

YELLOW_FILL = PatternFill(
    fill_type="solid",
    start_color="FFF59D",
    end_color="FFF59D"
)

RED_FILL = PatternFill(
    fill_type="solid",
    start_color="FFB6B6",
    end_color="FFB6B6"
)

GOLD_FILL = PatternFill(
    fill_type="solid",
    start_color="FFD700",
    end_color="FFD700"
)

def load_financial_ratios():

    conn = sqlite3.connect(DATABASE)

    df = pd.read_sql(
        "SELECT * FROM financial_ratios",
        conn
    )

    conn.close()

    return df

def load_peer_groups():

    conn = sqlite3.connect(DATABASE)

    df = pd.read_sql(
        "SELECT * FROM peer_groups",
        conn
    )

    conn.close()

    return df

def load_percentiles():

    conn = sqlite3.connect(DATABASE)

    df = pd.read_sql(
        "SELECT * FROM peer_percentiles",
        conn
    )

    conn.close()

    return df

def merge_data():

    ratios = load_financial_ratios()

    peers = load_peer_groups()

    merged = ratios.merge(

        peers,

        on="company_id",

        how="left"

    )

    return merged

def load_company_names():

    conn = sqlite3.connect(DATABASE)

    df = pd.read_sql(
        """
        SELECT
            id,
            company_name
        FROM companies
        """,
        conn
    )

    conn.close()

    return df

def pivot_percentiles():

    percentiles = load_percentiles()

    pivot = percentiles.pivot_table(

        index=[
            "company_id",
            "year"
        ],

        columns="metric",

        values="percentile_rank"

    ).reset_index()

    pivot.columns.name = None

    return pivot

PERCENTILE_COLUMNS = {

    "return_on_equity_pct":"roe_percentile",

    "return_on_capital_employed_pct":"roce_percentile",

    "net_profit_margin_pct":"npm_percentile",

    "debt_to_equity":"de_percentile",

    "free_cash_flow_cr":"fcf_percentile",

    "pat_cagr_5yr":"pat_percentile",

    "revenue_cagr_5yr":"revenue_percentile",

    "eps_cagr_5yr":"eps_percentile",

    "interest_coverage":"icr_percentile",

    "asset_turnover":"asset_percentile"

}

if __name__ == "__main__":

    ratios = load_financial_ratios()

    peers = load_peer_groups()

    percentiles = load_percentiles()

    print(ratios.shape)

    print(peers.shape)

    print(percentiles.shape)
    

    merged = merge_data()

    pivot = pivot_percentiles()
    pivot = pivot.rename(
    columns=PERCENTILE_COLUMNS
)

    merged = merged.merge(

    pivot,

    on=[
        "company_id",
        "year"
    ],

    how="left"

)

    print(merged.shape)

    print(merged.head())
    names = load_company_names()

    merged = merged.merge(

    names,

    left_on="company_id",

    right_on="id",

    how="left"

)
    print(merged.shape)

    print(merged.columns.tolist())
    print(

    merged[

[
"company_id",
"company_name",
"peer_group_name"

]

].head()

)
    os.makedirs(
    "output",
    exist_ok=True
)

    writer = pd.ExcelWriter(
    "output/peer_comparison.xlsx",
    engine="openpyxl"
)

    groups = sorted(
    merged["peer_group_name"]
    .dropna()
    .unique()
)

    for group in groups:

        sheet = merged[
        merged["peer_group_name"] == group
    ].copy()

        sheet = sheet.sort_values(
        "company_name"
    )

        sheet.to_excel(
        writer,
        sheet_name=group[:31],
        index=False
    )

        print(f"{group} exported.")

    writer.close()
    # ------------------------------------
# Apply Excel Formatting
# ------------------------------------

    wb = load_workbook("output/peer_comparison.xlsx")

    GREEN_FILL = PatternFill(
    fill_type="solid",
    start_color="90EE90",
    end_color="90EE90"
)

    YELLOW_FILL = PatternFill(
    fill_type="solid",
    start_color="FFF59D",
    end_color="FFF59D"
)

    RED_FILL = PatternFill(
    fill_type="solid",
    start_color="FFB6B6",
    end_color="FFB6B6"
)

    GOLD_FILL = PatternFill(
    fill_type="solid",
    start_color="FFD700",
    end_color="FFD700"
)

    percentile_columns = [
    "roe_percentile",
    "roce_percentile",
    "npm_percentile",
    "de_percentile",
    "fcf_percentile",
    "pat_percentile",
    "revenue_percentile",
    "eps_percentile",
    "icr_percentile",
    "asset_percentile"
]

    numeric_columns = [
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "debt_to_equity",
    "free_cash_flow_cr",
    "pat_cagr_5yr",
    "revenue_cagr_5yr",
    "eps_cagr_5yr",
    "interest_coverage",
    "asset_turnover"
]

    for ws in wb.worksheets:

        headers = {}

        for cell in ws[1]:
            headers[cell.value] = cell.column

    # -------------------------
    # Colour Percentiles
    # -------------------------

        for column in percentile_columns:

            if column not in headers:
                continue

            col = headers[column]

            for row in range(2, ws.max_row + 1):

               cell = ws.cell(row=row, column=col)

               if cell.value is None:
                   continue

               try:
                   value = float(cell.value)
               except:
                   continue

               if value >= 75:
                   cell.fill = GREEN_FILL

               elif value <= 25:
                   cell.fill = RED_FILL

               else:
                   cell.fill = YELLOW_FILL

    # -------------------------
    # Highlight Benchmark Row
    # -------------------------

        if "is_benchmark" in headers:

           benchmark_col = headers["is_benchmark"]

           for row in range(2, ws.max_row + 1):

                benchmark = ws.cell(
                row=row,
                column=benchmark_col
            ).value

                if benchmark == 1:

                    for col in range(1, ws.max_column + 1):

                        ws.cell(
                        row=row,
                        column=col
                    ).fill = GOLD_FILL

    # -------------------------
    # Median Summary Row
    # -------------------------
  
        summary_row = ws.max_row + 2

        ws.cell(
        row=summary_row,
        column=1
    ).value = "Peer Group Median"

        for column in numeric_columns:

            if column not in headers:
               continue

            excel_col = headers[column]

            values = []

            for row in range(2, ws.max_row + 1):

                value = ws.cell(
                row=row,
                column=excel_col
            ).value

                if isinstance(value, (int, float)):
                    values.append(value)

            if values:

                ws.cell(
                row=summary_row,
                column=excel_col
            ).value = round(
                pd.Series(values).median(),
                2
            )

    wb.save("output/peer_comparison.xlsx")

    print("\nFormatting Applied Successfully!")

    

